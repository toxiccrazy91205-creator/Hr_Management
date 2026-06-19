"""
views.py — Django views for the HR Recruitment System.

Views:
  • dashboard_view       — Upload resumes + enter job details → trigger AI pipeline
  • approval_view        — Human-in-the-loop review of shortlist & email drafts
  • approve_and_send_view — Resume the LangGraph to dispatch emails
  • success_view         — Confirmation page after emails are sent
"""

import os
import uuid
import json
import logging
import traceback

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponseBadRequest, HttpResponse, JsonResponse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone

from datetime import datetime
from dateutil import parser as dt_parser

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from core.models import JobPosting, Candidate, EmployeeProfile, CompanyPolicy, FAQ, PerformanceRecord, AttendanceLog

def is_hr_check(user):
    return user.is_authenticated and hasattr(user, 'employeeprofile') and user.employeeprofile.is_hr

def is_employee_check(user):
    return user.is_authenticated and hasattr(user, 'employeeprofile') and not user.employeeprofile.is_hr

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Helper: save uploaded files to MEDIA_ROOT/resumes/
# ──────────────────────────────────────────────────────────────────────────────

def _save_uploaded_files(files) -> list[str]:
    """Persist uploaded files to disk and return their absolute paths."""
    upload_dir = os.path.join(settings.MEDIA_ROOT, "resumes")
    os.makedirs(upload_dir, exist_ok=True)
    saved_paths = []
    for f in files:
        # Prefix with uuid to avoid collisions
        filename = f"{uuid.uuid4().hex[:8]}_{f.name}"
        dest = os.path.join(upload_dir, filename)
        with open(dest, "wb+") as fh:
            for chunk in f.chunks():
                fh.write(chunk)
        saved_paths.append(dest)
    return saved_paths


# ──────────────────────────────────────────────────────────────────────────────
# View 1: Dashboard — job form + resume upload
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def dashboard_view(request):
    """
    GET  → render the dashboard form (optionally pre-filled via ?resume=<job_id>).
    POST → accept job details + resumes, run the AI pipeline (steps 1-6),
           then redirect to the approval page.
    """
    if request.method == "POST":
        # ── Validate inputs ──────────────────────────────────────────────
        title = request.POST.get("job_title", "").strip()
        description = request.POST.get("job_description", "").strip()
        skills = request.POST.get("required_skills", "").strip()
        experience = request.POST.get("experience_years", "0").strip()
        files = request.FILES.getlist("resumes")
        resume_job_id = request.POST.get("resume_job_id", "").strip()

        if not title or not description or not skills:
            messages.error(request, "Please fill in all required fields.")
            return render(request, "dashboard.html", {"recent_jobs": JobPosting.objects.all()[:10]})

        if not files:
            messages.error(request, "Please upload at least one resume (PDF or DOCX).")
            return render(request, "dashboard.html", {"recent_jobs": JobPosting.objects.all()[:10]})

        try:
            experience_years = int(experience)
        except ValueError:
            experience_years = 0

        # ── Save files to disk ───────────────────────────────────────────
        saved_paths = _save_uploaded_files(files)

        # ── Create or reuse JobPosting record ────────────────────────────
        if resume_job_id:
            try:
                job = JobPosting.objects.get(pk=resume_job_id)
                job.title = title
                job.description = description
                job.required_skills = skills
                job.experience_years = experience_years
                job.status = "screening"
                job.save()
                # Clear old candidates from any previous failed run
                job.candidates.all().delete()
            except JobPosting.DoesNotExist:
                job = JobPosting.objects.create(
                    title=title,
                    description=description,
                    required_skills=skills,
                    experience_years=experience_years,
                    status="screening",
                )
        else:
            job = JobPosting.objects.create(
                title=title,
                description=description,
                required_skills=skills,
                experience_years=experience_years,
                status="screening",
            )

        # ── Run the LangGraph (steps 1-6, pauses before send_emails) ────
        thread_id = str(job.id)
        config = {"configurable": {"thread_id": thread_id}}

        initial_state = {
            "job_id": str(job.id),
            "job_title": title,
            "job_description": description,
            "required_skills": skills,
            "experience_years": experience_years,
            "uploaded_file_paths": saved_paths,
        }

        try:
            from ai_engine.graph import hr_graph

            # .invoke() runs until interrupt_before pauses the graph
            result = hr_graph.invoke(initial_state, config=config)

            # ── Persist shortlisted candidates into DB ───────────────────
            shortlisted = result.get("shortlisted_candidates", [])
            email_drafts = result.get("email_drafts", [])
            scheduled = result.get("scheduled_candidates", [])

            # Build lookup of email drafts by name
            draft_map = {d.get("name", ""): d for d in email_drafts}

            for candidate_data in shortlisted:
                name = candidate_data.get("name", "Unknown")
                email = candidate_data.get("email", "")
                score = candidate_data.get("match_score", 0)
                matched = candidate_data.get("matched_skills", "")
                slot = candidate_data.get("interview_slot", "TBD")
                source = candidate_data.get("source_file", "")

                # Get corresponding email draft
                draft_info = draft_map.get(name, {})
                subject = draft_info.get("subject", "")
                body = draft_info.get("body", "")
                drafted_email_text = f"Subject: {subject}\n\n{body}" if subject else ""

                Candidate.objects.create(
                    job=job,
                    name=name,
                    email=email if email else None,
                    match_score=float(score),
                    matched_skills=matched,
                    interview_slot=slot,
                    drafted_email=drafted_email_text,
                    resume_source=source,
                    status="email_sent",
                    email_sent=True,
                )

            # Save thread ID and update status
            job.langgraph_thread_id = thread_id
            job.status = "approved"
            job.save()

            messages.success(
                request,
                f"AI screening complete! {len(shortlisted)} candidate(s) shortlisted and emails have been sent."
            )
            return redirect("core:success", job_id=job.id)

        except Exception as exc:
            logger.exception("AI pipeline failed")
            job.status = "draft"
            job.save()
            messages.error(
                request,
                f"AI pipeline error: {exc}. Please check your API key and try again."
            )
            return render(request, "dashboard.html", {
                "recent_jobs": JobPosting.objects.all()[:10],
                "resume_job": job,
            })

    # GET request — check for ?resume=<job_id> to pre-fill form
    recent_jobs = JobPosting.objects.all()[:10]
    resume_job = None
    resume_id = request.GET.get("resume")
    if resume_id:
        try:
            resume_job = JobPosting.objects.get(pk=resume_id)
        except (JobPosting.DoesNotExist, ValueError):
            pass

    return render(request, "dashboard.html", {
        "recent_jobs": recent_jobs,
        "resume_job": resume_job,
    })


# ──────────────────────────────────────────────────────────────────────────────
# View 4: Success — final confirmation
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def success_view(request, job_id):
    """Display the final success page with sent-email summary."""
    job = get_object_or_404(JobPosting, pk=job_id)
    candidates = job.candidates.filter(email_sent=True)
    return render(request, "success.html", {
        "job": job,
        "candidates": candidates,
    })


# ──────────────────────────────────────────────────────────────────────────────
# View 5: Attendance Tracker — Upload timesheet
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def attendance_upload_view(request):
    """
    GET  → Render the attendance page with card tap device (all users)
           and optional timesheet upload + activity log (HR only).
    POST → (HR only) Save the uploaded Excel file, generate report, and store in session.
    """
    user = request.user
    is_hr = hasattr(user, 'employeeprofile') and user.employeeprofile.is_hr

    # Build the employee ID for the card tap device
    if is_hr:
        employee_id = f"HR-{user.username.upper()}"
    else:
        employee_id = f"EMP-{user.id:04d}"

    if request.method == "POST" and is_hr:
        uploaded_file = request.FILES.get("attendance_file")
        if not uploaded_file:
            messages.error(request, "Please upload an Excel file.")
            return render(request, "attendance.html", {
                "is_hr": is_hr,
                "employee_id": employee_id,
                "recent_logs": AttendanceLog.objects.all()[:50] if is_hr else [],
            })
            
        if not (uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls')):
            messages.error(request, "Invalid file format. Please upload .xlsx or .xls files only.")
            return render(request, "attendance.html", {
                "is_hr": is_hr,
                "employee_id": employee_id,
                "recent_logs": AttendanceLog.objects.all()[:50] if is_hr else [],
            })

        # Save temporarily
        upload_dir = os.path.join(settings.MEDIA_ROOT, "attendance")
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"{uuid.uuid4().hex[:8]}_{uploaded_file.name}"
        file_path = os.path.join(upload_dir, filename)
        
        with open(file_path, "wb+") as fh:
            for chunk in uploaded_file.chunks():
                fh.write(chunk)
                
        try:
            from ai_engine.attendance import generate_attendance_report
            reports = generate_attendance_report(file_path)
            
            # Store reports in session to pass to the report view
            request.session['attendance_reports'] = reports
            messages.success(request, "Timesheet processed successfully!")
            return redirect("core:attendance_report")
            
        except Exception as exc:
            logger.exception("Attendance processing failed")
            messages.error(request, f"Failed to process timesheet: {exc}")
            return render(request, "attendance.html", {
                "is_hr": is_hr,
                "employee_id": employee_id,
                "recent_logs": AttendanceLog.objects.all()[:50] if is_hr else [],
            })

    # GET request context
    context = {
        "is_hr": is_hr,
        "employee_id": employee_id,
        "recent_logs": AttendanceLog.objects.all()[:50] if is_hr else [],
    }
    return render(request, "attendance.html", context)


# ──────────────────────────────────────────────────────────────────────────────
# View 6: Attendance Tracker — View report
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def attendance_report_view(request):
    """Display the AI-generated attendance report from session."""
    reports = request.session.get('attendance_reports', [])
    return render(request, "attendance_report.html", {"reports": reports})


# ──────────────────────────────────────────────────────────────────────────────
# View 7: Attendance Tracker — Download report as PDF
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def attendance_download_pdf_view(request):
    """Download the attendance report from session as a PDF file."""
    reports = request.session.get('attendance_reports', [])
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    elements.append(Paragraph("AI Attendance Report", title_style))
    elements.append(Spacer(1, 20))
    
    data = [['Employee Name', 'Total Hours', 'Leaves Taken', 'AI Score', 'Agent Feedback']]
    
    for report in reports:
        data.append([
            report.get('name', ''),
            str(report.get('total_hours', '')),
            str(report.get('leaves_taken', '')),
            str(report.get('ai_score', '')),
            Paragraph(str(report.get('ai_feedback', '')), styles['Normal'])
        ])
        
    table = Table(data, colWidths=[120, 80, 80, 70, 370])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#0a0e1a')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


# ──────────────────────────────────────────────────────────────────────────────
# View 8: Authentication Views (Login, Signup, Logout)
# ──────────────────────────────────────────────────────────────────────────────

from django.views.decorators.cache import never_cache

@never_cache
def custom_login_view(request):
    if request.user.is_authenticated:
        if hasattr(request.user, 'employeeprofile') and request.user.employeeprofile.is_hr:
            return redirect("core:dashboard")
        return redirect("core:employee_chat")
        
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            if hasattr(user, 'employeeprofile') and user.employeeprofile.is_hr:
                return redirect("core:dashboard")
            return redirect("core:employee_chat")
    else:
        form = AuthenticationForm()
    return render(request, "login.html", {"form": form})

def signup_view(request):
    if request.method == "POST":
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Default new signups to Employee
            EmployeeProfile.objects.create(
                user=user,
                is_hr=False,
                department="General",
                salary=50000.00,
                leave_balance=10,
                attendance_score=100
            )
            login(request, user)
            return redirect("core:employee_chat")
    else:
        form = UserCreationForm()
    return render(request, "signup.html", {"form": form})

def custom_logout_view(request):
    logout(request)
    return redirect("core:login")


# ──────────────────────────────────────────────────────────────────────────────
# View 9: HR Management (CRUD for Policies & FAQs)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def hr_data_management_view(request):
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_policy":
            policy_id = request.POST.get("policy_id")
            if policy_id:
                CompanyPolicy.objects.filter(id=policy_id).update(
                    title=request.POST.get("title"),
                    content=request.POST.get("content")
                )
                messages.success(request, "Policy updated successfully.")
            else:
                CompanyPolicy.objects.create(
                    title=request.POST.get("title"),
                    content=request.POST.get("content")
                )
                messages.success(request, "Policy added successfully.")
        elif action == "delete_policy":
            policy_id = request.POST.get("policy_id")
            if policy_id:
                CompanyPolicy.objects.filter(id=policy_id).delete()
                messages.success(request, "Policy deleted successfully.")
        elif action == "add_faq":
            faq_id = request.POST.get("faq_id")
            if faq_id:
                FAQ.objects.filter(id=faq_id).update(
                    question=request.POST.get("question"),
                    answer=request.POST.get("answer")
                )
                messages.success(request, "FAQ updated successfully.")
            else:
                FAQ.objects.create(
                    question=request.POST.get("question"),
                    answer=request.POST.get("answer")
                )
                messages.success(request, "FAQ added successfully.")
        elif action == "delete_faq":
            faq_id = request.POST.get("faq_id")
            if faq_id:
                FAQ.objects.filter(id=faq_id).delete()
                messages.success(request, "FAQ deleted successfully.")
        return redirect("core:hr_manage_data")
        
    policies = CompanyPolicy.objects.all()
    faqs = FAQ.objects.all()
    employees = EmployeeProfile.objects.filter(is_hr=False)
    
    return render(request, "hr_manage_data.html", {
        "policies": policies,
        "faqs": faqs,
        "employees": employees
    })


# ──────────────────────────────────────────────────────────────────────────────
# View 10: Employee Chatbot
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_employee_check)
def employee_chat_view(request):
    return render(request, "employee_chat.html")

@login_required
@user_passes_test(is_employee_check)
def api_chat_view(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            message = data.get("message", "")
            
            from ai_engine.support_agent import generate_support_response
            response_text = generate_support_response(message, request.user.employeeprofile)
            
            return JsonResponse({"response": response_text})
        except Exception as e:
            logger.exception("Chat API Error")
            return JsonResponse({"error": str(e)}, status=500)
    return JsonResponse({"error": "Invalid method"}, status=405)


# ──────────────────────────────────────────────────────────────────────────────
# View 11: Performance Review Dashboard (Bulk Upload)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def hr_performance_dashboard_view(request):
    """Handles global Excel upload for bulk Performance Reviews."""
    if request.method == "POST":
        review_period = request.POST.get("review_period")
        excel_file = request.FILES.get("performance_excel")
        
        try:
            import pandas as pd
            if not excel_file:
                raise ValueError("No Excel file uploaded.")
                
            df = pd.read_excel(excel_file)
            
            # Flexible identifier column matching
            identifier_col = None
            is_id = False
            for col in df.columns:
                col_lower = col.lower()
                if 'email' in col_lower or 'username' in col_lower or 'name' in col_lower:
                    identifier_col = col
                    break
                elif 'id' == col_lower or 'employee id' in col_lower:
                    identifier_col = col
                    is_id = True
                    break
            
            if not identifier_col:
                raise ValueError("Excel file must contain an 'Employee Name', 'Username', 'Email', or 'Employee ID' column to identify employees.")
            
            processed_record_ids = []
            
            for index, row in df.iterrows():
                identifier = str(row[identifier_col]).strip()
                
                # Try to find the employee
                employee = None
                try:
                    if is_id and identifier.replace('.','',1).isdigit():
                        employee = EmployeeProfile.objects.get(id=int(float(identifier)), is_hr=False)
                    elif '@' in identifier:
                        employee = EmployeeProfile.objects.get(user__email=identifier, is_hr=False)
                    else:
                        employee = EmployeeProfile.objects.get(user__username=identifier, is_hr=False)
                except EmployeeProfile.DoesNotExist:
                    pass # We will auto-create below
                
                # Auto-create the employee if they don't exist in the DB (great for testing arbitrary Excel files)
                if not employee:
                    from django.contrib.auth.models import User
                    
                    # Try to get an actual name if the identifier was an ID
                    display_name = str(identifier)
                    if is_id:
                        for col in row.index:
                            if 'name' in col.lower():
                                display_name = str(row[col])
                                break
                                
                    safe_username = str(display_name).replace(" ", "_").lower()
                    if not safe_username:
                        continue
                        
                    # Handle case where user might already exist but no EmployeeProfile
                    try:
                        user = User.objects.get(username=safe_username)
                    except User.DoesNotExist:
                        user = User.objects.create(username=safe_username, email=f"{safe_username}@demo.com", first_name=str(display_name))
                        
                    employee, _ = EmployeeProfile.objects.get_or_create(
                        user=user,
                        defaults={'is_hr': False, 'department': 'General'}
                    )
                
                # Process with LLM
                from ai_engine.performance_agent import generate_performance_review
                # Convert row to dict
                row_dict = row.to_dict()
                row_data_str = ", ".join([f"{k}: {v}" for k, v in row_dict.items()])
                
                result = generate_performance_review(
                    employee_profile=employee,
                    row_data=row_data_str
                )
                
                # Save the record
                record = PerformanceRecord.objects.create(
                    employee=employee,
                    review_period=review_period,
                    kpi_data=row_data_str, # Store row data here
                    task_completion_rate=0.0, # Deprecated
                    manager_feedback="See Excel Data", # Deprecated
                    performance_score=result.get("performance_score"),
                    kpi_achievement_percent=result.get("kpi_achievement_percent"),
                    strengths=result.get("strengths"),
                    areas_for_improvement=result.get("areas_for_improvement"),
                    recommended_actions=result.get("recommended_actions"),
                )
                processed_record_ids.append(record.id)
            
            if not processed_record_ids:
                raise ValueError("No matching employees found in the Excel file.")
                
            # Store processed IDs in session to display them on the report page
            request.session['latest_performance_records'] = processed_record_ids
            
            messages.success(request, f"Successfully generated {len(processed_record_ids)} performance reviews!")
            return redirect("core:view_performance_report")
            
        except Exception as e:
            logger.exception("Performance bulk generation failed")
            messages.error(request, f"Failed to generate reviews: {str(e)}")
            return redirect("core:hr_performance_dashboard")

    return render(request, "hr_performance_dashboard.html")

# ──────────────────────────────────────────────────────────────────────────────
# View 12: View Performance Report
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def view_performance_report_view(request):
    """Displays the AI-generated structured performance reports from the bulk upload."""
    record_ids = request.session.get('latest_performance_records', [])
    records = PerformanceRecord.objects.filter(id__in=record_ids)
    
    if not records.exists():
        messages.warning(request, "No recent reports to display.")
        return redirect("core:hr_performance_dashboard")
        
    return render(request, "performance_report.html", {"records": records})

# ──────────────────────────────────────────────────────────────────────────────
# View 13: Download Performance PDF
# ──────────────────────────────────────────────────────────────────────────────
import io
from django.http import HttpResponse

@login_required
@user_passes_test(is_hr_check)
def performance_download_pdf_view(request):
    """Generates a PDF of all recently processed performance reports."""
    record_ids = request.session.get('latest_performance_records', [])
    records = PerformanceRecord.objects.filter(id__in=record_ids)
    
    if not records.exists():
        messages.error(request, "No reports available for download.")
        return redirect("core:hr_performance_dashboard")
        
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = styles['Heading1']
    sub_style = styles['Heading2']
    normal_style = styles['Normal']
    
    elements.append(Paragraph("Performance Review Reports (Bulk Generated)", title_style))
    elements.append(Spacer(1, 20))
    
    for record in records:
        elements.append(Paragraph(f"Employee: {record.employee.user.username} ({record.employee.user.email})", sub_style))
        elements.append(Paragraph(f"Review Period: {record.review_period}", normal_style))
        elements.append(Paragraph(f"AI Performance Score: {record.performance_score}/100", normal_style))
        elements.append(Paragraph(f"KPI Achievement: {record.kpi_achievement_percent}%", normal_style))
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Strengths:", styles['Heading4']))
        elements.append(Paragraph(record.strengths.replace('\n', '<br/>'), normal_style))
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Areas for Improvement:", styles['Heading4']))
        elements.append(Paragraph(record.areas_for_improvement.replace('\n', '<br/>'), normal_style))
        
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Recommended Actions:", styles['Heading4']))
        elements.append(Paragraph(record.recommended_actions.replace('\n', '<br/>'), normal_style))
        
        elements.append(Spacer(1, 30))
        
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="performance_reports.pdf"'
    return response


# ──────────────────────────────────────────────────────────────────────────────
# View 14: Hardware Tap API — Fingerprint / ID Card Attendance Webhook
# ──────────────────────────────────────────────────────────────────────────────


# ──────────────────────────────────────────────────────────────────────────────
# View 15: Export Attendance Logs to Excel (HR Only)
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@user_passes_test(is_hr_check)
def attendance_export_excel_view(request):
    """Export all attendance tap logs to an Excel (.xlsx) file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback: generate a CSV instead if openpyxl is not installed
        import csv
        logs = AttendanceLog.objects.all()
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_logs.csv"'
        writer = csv.writer(response)
        writer.writerow(['Employee ID', 'Timestamp', 'Tap Type', 'Device Type', 'Location'])
        for log in logs:
            writer.writerow([
                log.employee_id,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.get_tap_type_display(),
                log.device_type,
                log.location,
            ])
        return response

    logs = AttendanceLog.objects.all().order_by('-timestamp')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"

    # Styling
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    in_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    out_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')

    # Headers
    headers = ['#', 'Employee ID', 'Date', 'Time', 'Tap Type', 'Device Type', 'Location']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_num, log in enumerate(logs, 2):
        tap_display = 'CHECK-IN' if log.tap_type == 'IN' else 'CHECK-OUT'
        row_data = [
            row_num - 1,
            log.employee_id,
            log.timestamp.strftime('%Y-%m-%d'),
            log.timestamp.strftime('%H:%M:%S'),
            tap_display,
            log.device_type,
            log.location,
        ]
        row_fill = in_fill if log.tap_type == 'IN' else out_fill
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center' if col_num in [1, 5] else 'left', vertical='center')

    # Column widths
    col_widths = [6, 18, 14, 12, 14, 18, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

@csrf_exempt
def hardware_tap_api(request):
    """
    POST-only endpoint that simulates receiving a tap from physical hardware
    (RFID scanner, NFC reader, biometric device, etc.).

    Logic:
      • Count today's taps for the given employee.
      • Even count (0, 2, 4 …) → this tap is "IN"
      • Odd count  (1, 3, 5 …) → this tap is "OUT"

    Expected JSON payload:
        {
            "employeeId": "EMP-8472",
            "timestamp":  "2026-06-19T12:00:00Z",
            "deviceType": "rfid_scanner",
            "location":   "Main Entrance Gate"
        }
    """
    if request.method != "POST":
        return JsonResponse({"error": "Only POST requests are accepted."}, status=405)

    try:
        data = json.loads(request.body)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON payload."}, status=400)

    employee_id = data.get("employeeId", "").strip()
    timestamp_str = data.get("timestamp", "").strip()
    device_type = data.get("deviceType", "rfid_scanner").strip()
    location = data.get("location", "Main Entrance").strip()

    if not employee_id:
        return JsonResponse({"error": "Missing required field: employeeId."}, status=400)

    # Parse the timestamp (fall back to current time if missing / invalid)
    if timestamp_str:
        try:
            tap_time = dt_parser.isoparse(timestamp_str)
            # Make timezone-aware if naive
            if tap_time.tzinfo is None:
                tap_time = timezone.make_aware(tap_time)
        except (ValueError, OverflowError):
            tap_time = timezone.now()
    else:
        tap_time = timezone.now()

    # Count today's existing taps for this employee
    today_start = tap_time.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = tap_time.replace(hour=23, minute=59, second=59, microsecond=999999)

    today_tap_count = AttendanceLog.objects.filter(
        employee_id=employee_id,
        timestamp__range=(today_start, today_end),
    ).count()

    # Even → IN, Odd → OUT
    tap_type = "IN" if today_tap_count % 2 == 0 else "OUT"

    log_entry = AttendanceLog.objects.create(
        employee_id=employee_id,
        timestamp=tap_time,
        device_type=device_type,
        location=location,
        tap_type=tap_type,
    )

    return JsonResponse({
        "status": "success",
        "message": f"Tap recorded: {tap_type}",
        "data": {
            "id": log_entry.id,
            "employeeId": log_entry.employee_id,
            "tapType": log_entry.tap_type,
            "tapNumber": today_tap_count + 1,
            "timestamp": log_entry.timestamp.isoformat(),
            "deviceType": log_entry.device_type,
            "location": log_entry.location,
        }
    }, status=201)
